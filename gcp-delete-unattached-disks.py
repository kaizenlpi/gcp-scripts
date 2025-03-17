"""The script ingests an excel spreadsheet to parse the list of disks to delete that are unattached."""
import subprocess
import json
from openpyxl import load_workbook

# Set GCP project
gcp_project = "MY-PROJECT-ID"  # Replace with your actual GCP project ID

# Define regions and their possible zones
region_zones = {
    "northamerica-northeast1": ["northamerica-northeast1-a", "northamerica-northeast1-b", "northamerica-northeast1-c"],
    "northamerica-northeast2": ["northamerica-northeast2-a", "northamerica-northeast2-b", "northamerica-northeast2-c"],
    "us-east4": ["us-east4-a", "us-east4-b", "us-east4-c"]
}

# Load the Excel file containing the disk names
wb = load_workbook("CRE-PCE-GCP-Unattached-Volumes.xlsx")
ws = wb["Unattached Volumes - CRE PCE"]
column = ws["G"]  # Adjust column as needed

# Extract disk names from the spreadsheet
disk_names = [cell.value for cell in column if cell.value]

# Function to get the region of a disk by checking each zone
def get_disk_region(disk_name):
    """Fetch the correct region of the given disk by checking each zone."""
    for region, zones in region_zones.items():
        for zone in zones:
            command = f"gcloud compute disks describe {disk_name} --project={gcp_project} --zone={zone} --format=json"
            result = subprocess.run(command, shell=True, capture_output=True, text=True)

            if result.returncode == 0:
                print(f"Disk {disk_name} found in region {region}, zone {zone}.")
                return region
            else:
                # If disk is not found in this zone, try the next zone
                print(f"Disk {disk_name} not found in zone {zone}. Trying next zone...")

    # If disk is not found in any zone, return None
    print(f"Disk {disk_name} not found in any zone.")
    return None

# Loop through each disk and attempt deletion in all zones of its region
for disk in disk_names:
    print(f"Fetching region for disk: {disk}")
    region = get_disk_region(disk)

    if not region:
        print(f"Skipping {disk}: Unable to determine region.")
        continue

    print(f"Attempting to delete disk {disk} from region {region}")

    # If the region is known, loop through its zones
    if region in region_zones:
        for zone in region_zones[region]:
            print(f"Trying to delete {disk} in zone: {zone}")

            command = f"gcloud compute disks delete {disk} --project={gcp_project} --zone={zone} --quiet"
            result = subprocess.run(command, shell=True, capture_output=True, text=True)

            print("STDOUT:", result.stdout)
            print("STDERR:", result.stderr)

            if result.returncode == 0:
                print(f"✅ Disk {disk} deleted successfully in zone {zone}.")
                break  # Stop trying after a successful deletion
            else:
                print(f"❌ Failed to delete {disk} in zone {zone}. Trying next zone...")

        else:
            print(f"⚠️ Disk {disk} could not be deleted in any zone of {region}.")
    else:
        print(f"⚠️ No zones found for region {region}. Skipping disk {disk}.")

print("All disks processed.")
